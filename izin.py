import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog, Menu, Toplevel, Label, Entry, Button, ttk
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
import shutil
import win32print
import win32api
import json
import subprocess
import sys

CONFIG_FILE = "config.txt"
SETTINGS_FILE = "settings.json"

DEFAULT_SETTINGS = {
    "Tarih": {"cells": "G1,B19,F19,F22,B22", "enabled": True},
    "Kisim_Sefi": {"cells": "B21,F5", "enabled": True},
    "Adi_Soyadi": {"cells": "D3", "enabled": True},
    "Gorevi": {"cells": "D4", "enabled": True},
    "Talep_Edilen_Izin_Gun_Sayisi": {"cells": "D6", "enabled": True},
    "Izin_Baslangic_Tarihi": {"cells": "D7", "enabled": True},
    "Izin_Bitis_Tarihi": {"cells": "G7", "enabled": True},
    "Izinli_Iken_Yerine_Bakacak_Kisi": {"cells": "D8", "enabled": True},
    "Kontrol_Eden": {"cells": "B18", "enabled": True},
    "Kontrol_Eden_Adi": "EMRAH GÜMRÜK"
}

AZURE_THEME_PATH = "azure.tcl"  # azure.tcl dosyasının bulunduğu yol

def kisi_listesi_yukle():
    try:
        with open("kisi.txt", "r", encoding="utf-8") as file:
            kisiler = [line.strip().split(":") for line in file if line.strip()]
        return kisiler
    except FileNotFoundError:
        return []

def parse_date(date_string):
    return datetime.strptime(date_string, "%d.%m.%Y")

def calculate_days(start_date, end_date):
    start = parse_date(start_date)
    end = parse_date(end_date) - timedelta(days=1)  # bitiş gününü sil
    total_days = (end - start).days + 1  # ilk günü say
    sunday_count = sum(1 for i in range(total_days) if (start + timedelta(days=i)).weekday() == 6)
    requested_days = total_days - sunday_count  # toplam izinden pazarı sil
    return requested_days, sunday_count

def turkish_upper(text):
    return text.replace('i', 'İ').replace('ı', 'I').replace('ş', 'Ş').replace('ç', 'Ç').replace('ğ', 'Ğ').replace('ü', 'Ü').replace('ö', 'Ö').upper()

class IzinFormuApp:
    def __init__(self, master):
        self.master = master
        master.title("İzin Formu")
        self.master.tk.call("source", AZURE_THEME_PATH)
        self.master.tk.call("set_theme", "dark")
        self.master.configure(background="#333333")  # Pencere arka plan rengini ayarlayın

        self.master.option_add('*TCombobox*Listbox.background', '#333333')
        self.master.option_add('*TCombobox*Listbox.foreground', 'white')
        self.master.option_add('*TCombobox*Listbox.selectBackground', '#444444')
        self.master.option_add('*TCombobox*Listbox.selectForeground', 'white')
        
        self.style = ttk.Style()
        self.style.configure('TButton', background='#333333', foreground='white')
        self.style.map('TButton', background=[('active', '#444444')])
        self.style.configure('TLabel', background='#333333', foreground='white')
        self.style.configure('TCheckbutton', background='#333333', foreground='white')
        self.style.configure('TEntry', fieldbackground='#333333', foreground='white', background='#333333')

        self.kisiler = kisi_listesi_yukle()
        self.default_save_path = self.load_config()
        self.settings = self.load_settings()
        self.setup_ui()

        # Tema ayarını self.setup_ui çağrısından sonra yapıyoruz
        self.change_theme("azure-dark")

    def setup_ui(self):
        self.menu = Menu(self.master, background="#333333", foreground="white", activebackground="#444444", activeforeground="white")
        self.master.config(menu=self.menu)
        
        self.settings_menu = Menu(self.menu, tearoff=0, background="#333333", foreground="white", activebackground="#444444", activeforeground="white")
        self.menu.add_cascade(label="Settings", menu=self.settings_menu)
        self.settings_menu.add_command(label="Ayarları Düzenle", command=self.open_settings_window)
        self.settings_menu.add_command(label="Kişileri Düzenle", command=self.edit_kisiler)

        self.labels = ["Tarih", "Adi_Soyadi", "Gorevi", "Kisim_Sefi", "Talep_Edilen_Izin_Gun_Sayisi", "Izin_Baslangic_Tarihi", "Izin_Bitis_Tarihi", "Izinli_Iken_Yerine_Bakacak_Kisi"]
        self.entries = {}
        self.checkbuttons = {}
        self.checkbutton_vars = {}
        self.original_values = {}
        today = datetime.now().strftime("%d.%m.%Y")
        for idx, label in enumerate(self.labels):
            ttk.Label(self.master, text=label.replace('_', ' ')).grid(row=idx, column=0, sticky="w")
            if label in ["Tarih", "Izin_Baslangic_Tarihi", "Izin_Bitis_Tarihi"]:
                entry_var = tk.StringVar(value=today)
                entry = DateEntry(self.master, textvariable=entry_var, date_pattern='dd.mm.yyyy')
                entry.bind("<<DateEntrySelected>>", lambda event, var=entry_var: self.update_leave_days())
            else:
                entry_var = tk.StringVar()
                entry = ttk.Entry(self.master, textvariable=entry_var)
            entry.grid(row=idx, column=1, sticky="ew")
            self.entries[label] = entry_var
            if label == "Adi_Soyadi":
                entry.bind("<Return>", self.on_ad_soyad_enter)
            
            checkbutton_var = tk.BooleanVar(value=self.settings[label]["enabled"])
            checkbutton = ttk.Checkbutton(self.master, variable=checkbutton_var)
            checkbutton.grid(row=idx, column=2, sticky="w")
            self.checkbuttons[label] = checkbutton
            self.checkbutton_vars[label] = checkbutton_var

        self.belirsiz_var = tk.IntVar()
        self.belirsiz_checkbox = ttk.Checkbutton(self.master, text="Ucu Açık İzin", variable=self.belirsiz_var, command=self.toggle_belirsiz)
        self.belirsiz_checkbox.grid(row=len(self.labels), column=0, sticky="w")

        self.yarim_gun_var = tk.IntVar()
        self.yarim_gun_checkbox = ttk.Checkbutton(self.master, text="Yarım Gün İzin", variable=self.yarim_gun_var, command=self.toggle_yarim_gun)
        self.yarim_gun_checkbox.grid(row=len(self.labels) + 1, column=0, sticky="w")

        self.desktop_copy_var = tk.IntVar()
        self.desktop_copy_checkbox = ttk.Checkbutton(self.master, text="Masaüstüne Kopyala", variable=self.desktop_copy_var)
        self.desktop_copy_checkbox.grid(row=len(self.labels) + 2, column=0, sticky="w")

        ttk.Button(self.master, text="Yeniden Başlat", command=self.restart_program).grid(row=len(self.labels) + 3, column=0, columnspan=2, sticky="ew")
        ttk.Button(self.master, text="Kaydet", command=self.kaydet).grid(row=len(self.labels) + 4, column=0, columnspan=2, sticky="ew")
        ttk.Button(self.master, text="Kaydet ve Yazdır", command=self.yazdir).grid(row=len(self.labels) + 5, column=0, columnspan=2, sticky="ew")
        
        ttk.Button(self.master, text="Dosya Konumu Seç", command=self.select_save_location).grid(row=len(self.labels) + 6, column=0, columnspan=2, sticky="ew")

        self.save_path_label = ttk.Label(self.master, text=self.default_save_path, foreground="white", background="#333333")
        self.save_path_label.grid(row=len(self.labels) + 7, column=0, columnspan=2, sticky="ew")

        self.copyright_label = ttk.Label(self.master, text="www.emrahgumruk.com.tr", foreground="white", background="#333333")
        self.copyright_label.grid(row=len(self.labels) + 8, column=0, columnspan=2, sticky="ew")

        self.update_leave_days()  # Program başlatıldığında izin gün sayısını hesapla

    def change_theme(self, theme):
        self.style.theme_use(theme)

    def load_azure_theme(self, path):
        try:
            self.master.tk.call("source", path)
            self.style.theme_use("azure-dark")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load theme: {e}")

    def on_ad_soyad_enter(self, event=None):
        ad_soyad = turkish_upper(self.entries["Adi_Soyadi"].get())
        eslesen_kisiler = [kisi for kisi in self.kisiler if ad_soyad == turkish_upper(kisi[0].split()[0])]
        
        if not eslesen_kisiler:
            eslesen_kisiler = [kisi for kisi in self.kisiler if ad_soyad in turkish_upper(kisi[0])]
        
        if eslesen_kisiler:
            secilen = eslesen_kisiler[0]
            ad_soyad, gorevi, kisim_sefi = secilen
            self.entries["Adi_Soyadi"].set(ad_soyad)
            self.entries["Gorevi"].set(gorevi)
            self.entries["Kisim_Sefi"].set(kisim_sefi)
        else:
            messagebox.showerror("Hata", "Eşleşen kişi bulunamadı.")

    def update_leave_days(self, event=None):
        if self.belirsiz_var.get() == 1:
            self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set("BELİRSİZ")
            self.entries["Izin_Bitis_Tarihi"].set("BELİRSİZ")
        elif self.yarim_gun_var.get() == 1:
            self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set("0.5")
        else:
            baslangic_tarihi = self.entries["Izin_Baslangic_Tarihi"].get()
            bitis_tarihi = self.entries["Izin_Bitis_Tarihi"].get()
            if baslangic_tarihi and bitis_tarihi:
                requested_days, sunday_count = calculate_days(baslangic_tarihi, bitis_tarihi)
                talep_edilen_izin = f"{requested_days}" + (f"+{sunday_count}" if sunday_count else "")
                self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set(talep_edilen_izin)
            else:
                self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set("")

    def toggle_belirsiz(self):
        if self.belirsiz_var.get() == 1:
            self.original_values["Talep_Edilen_Izin_Gun_Sayisi"] = self.entries["Talep_Edilen_Izin_Gun_Sayisi"].get()
            self.original_values["Izin_Bitis_Tarihi"] = self.entries["Izin_Bitis_Tarihi"].get()
            self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set("BELİRSİZ")
            self.entries["Izin_Bitis_Tarihi"].set("BELİRSİZ")
        else:
            if "Talep_Edilen_Izin_Gun_Sayisi" in self.original_values:
                self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set(self.original_values["Talep_Edilen_Izin_Gun_Sayisi"])
            if "Izin_Bitis_Tarihi" in self.original_values:
                self.entries["Izin_Bitis_Tarihi"].set(self.original_values["Izin_Bitis_Tarihi"])
            self.update_leave_days()

    def toggle_yarim_gun(self):
        if self.yarim_gun_var.get() == 1:
            self.original_values["Talep_Edilen_Izin_Gun_Sayisi"] = self.entries["Talep_Edilen_Izin_Gun_Sayisi"].get()
            self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set("0.5")
        else:
            if "Talep_Edilen_Izin_Gun_Sayisi" in self.original_values:
                self.entries["Talep_Edilen_Izin_Gun_Sayisi"].set(self.original_values["Talep_Edilen_Izin_Gun_Sayisi"])
            self.update_leave_days()

    def select_save_location(self):
        selected_directory = filedialog.askdirectory()
        if selected_directory:
            self.default_save_path = selected_directory
            self.save_path_label.config(text=self.default_save_path)  # Seçilen dosya konumunu güncelle
            self.save_config()

    def save_config(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as file:
            file.write(self.default_save_path)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as file:
                return file.read().strip()
        return os.path.expanduser("~/Desktop")  # Varsayılan olarak masaüstü kullanılıyor

    def kaydet(self):
        dosya_adi = "izin2.xlsx" if self.belirsiz_var.get() == 1 else "izin.xlsx"
        if not os.path.exists(dosya_adi):
            wb = Workbook()
        else:
            wb = load_workbook(dosya_adi)
        ws = wb.active

        # Verilerin girildiği yerler
        for label, entry_var in self.entries.items():
            if self.checkbutton_vars[label].get():
                cells = self.settings[label]["cells"].split(',')
                for cell in cells:
                    ws[cell] = entry_var.get().upper()
        
        kontrol_eden_cells = self.settings["Kontrol_Eden"]["cells"].split(',')
        for cell in kontrol_eden_cells:
            ws[cell] = self.settings["Kontrol_Eden_Adi"]

        save_path = os.path.join(self.default_save_path, dosya_adi)
        wb.save(save_path)
        print(f"Excel dosyası '{save_path}' olarak kaydedildi.")

        # Yeni dosya adını oluşturma
        adi_soyadi = self.entries["Adi_Soyadi"].get().replace(" ", "_").upper()
        baslangic_tarihi = self.entries["Izin_Baslangic_Tarihi"].get().replace(".", "-")
        bitis_tarihi = self.entries["Izin_Bitis_Tarihi"].get().replace(".", "-")
        yeni_dosya_adi = f"{adi_soyadi}_{baslangic_tarihi}_{bitis_tarihi}.xlsx"

        # Dosyayı seçilen konuma kopyalama ve adını değiştirme
        full_save_path = os.path.join(self.default_save_path, yeni_dosya_adi)
        shutil.copy(save_path, full_save_path)
        print(f"Kopyalanan dosya '{full_save_path}' olarak kaydedildi.")

        # Ekstra klasör oluşturma ve dosyayı bu klasöre kopyalama
        ekstra_klasor_path = os.path.join(self.default_save_path, "Ekstra_Klasor")
        os.makedirs(ekstra_klasor_path, exist_ok=True)
        ekstra_dosya_path = os.path.join(ekstra_klasor_path, yeni_dosya_adi)
        shutil.copy(full_save_path, ekstra_dosya_path)
        print(f"Dosya ekstra klasöre '{ekstra_dosya_path}' olarak kopyalandı.")

        # Masaüstüne kopyalama
        if self.desktop_copy_var.get() == 1:
            desktop_path = os.path.join(os.path.expanduser("~/Desktop"), yeni_dosya_adi)
            shutil.copy(full_save_path, desktop_path)
            print(f"Kopyalanan dosya masaüstüne '{desktop_path}' olarak kaydedildi.")

        return full_save_path

    def yazdir(self):
        dosya_adi = self.kaydet()
        if dosya_adi:
            # Kaç adet yazdırılacağını sor
            adet = simpledialog.askinteger("Yazdırma Adedi", "Kaç adet yazdırmak istiyorsunuz?", initialvalue=1, minvalue=1)
            if adet:
                for _ in range(adet):
                    win32api.ShellExecute(0, "print", dosya_adi, None, ".", 0)
                print(f"{adet} adet yazdırıldı.")

    def restart_program(self):
        python = sys.executable
        os.execl(python, python, *sys.argv)

    def open_settings_window(self):
        settings_window = Toplevel(self.master)
        settings_window.title("Ayarları Düzenle")
        settings_window.configure(background="#333333")

        for idx, label in enumerate(self.labels):
            Label(settings_window, text=label.replace('_', ' ') + ":", background="#333333", foreground="white").grid(row=idx, column=0, sticky="w")
            cells_entry = Entry(settings_window, background="#444444", foreground="white", insertbackground="white")
            cells_entry.insert(0, self.settings[label]["cells"])
            cells_entry.grid(row=idx, column=1, sticky="ew")
            
            enabled_var = tk.BooleanVar(value=self.settings[label]["enabled"])
            enabled_checkbutton = ttk.Checkbutton(settings_window, variable=enabled_var)
            enabled_checkbutton.grid(row=idx, column=2, sticky="w")

            self.entries[label + "_entry"] = cells_entry
            self.checkbutton_vars[label + "_enabled"] = enabled_var

        kontrol_eden_adi_entry = Entry(settings_window, background="#444444", foreground="white", insertbackground="white")
        kontrol_eden_adi_entry.insert(0, self.settings["Kontrol_Eden_Adi"])
        kontrol_eden_adi_entry.grid(row=len(self.labels), column=1, sticky="ew")
        Label(settings_window, text="Kontrol Eden Adı:", background="#333333", foreground="white").grid(row=len(self.labels), column=0, sticky="w")

        Button(settings_window, text="Kaydet", command=lambda: self.save_settings(
            {label: {"cells": self.entries[label + "_entry"].get(), "enabled": self.checkbutton_vars[label + "_enabled"].get()} for label in self.labels},
            kontrol_eden_adi_entry.get()
        )).grid(row=len(self.labels) + 1, column=0, columnspan=3, sticky="ew")

    def save_settings(self, settings, kontrol_eden_adi):
        self.settings = settings
        self.settings["Kontrol_Eden_Adi"] = kontrol_eden_adi
        with open(SETTINGS_FILE, "w", encoding="utf-8") as file:
            json.dump(self.settings, file, ensure_ascii=False, indent=4)
        messagebox.showinfo("Bilgi", "Ayarlar kaydedildi")

    def load_settings(self):
        try:
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, "r", encoding="utf-8") as file:
                    return json.load(file)
        except json.JSONDecodeError:
            messagebox.showerror("Hata", "Ayarlar dosyası bozuk. Varsayılan ayarlar kullanılacak.")
        return DEFAULT_SETTINGS.copy()

    def edit_kisiler(self):
        kisi_file_path = os.path.join(os.getcwd(), "kisi.txt")
        if os.path.exists(kisi_file_path):
            if os.name == 'nt':  # Windows
                os.startfile(kisi_file_path)
            elif os.name == 'posix':  # macOS
                subprocess.call(('open', kisi_file_path))
            else:  # Linux and others
                subprocess.call(('xdg-open', kisi_file_path))
        else:
            messagebox.showerror("Hata", "kisi.txt dosyası bulunamadı.")

root = tk.Tk()
app = IzinFormuApp(root)
root.mainloop()
